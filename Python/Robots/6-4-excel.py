import openpyxl
site = '/Users/guojufeng/Documents/GitHub/JuFengGuo/Python/Robots/'

# 写入的代码
wb = openpyxl.Workbook() # 生成工作簿
sheet = wb.active # 获取当前活跃的工作表

sheet.title = 'myOneSheet' # 修改工作表的名称
sheet['A1'] = '姓名' # 单个单元格的颜色设置
sheet['B1'] = '年龄'
sheet['C1'] = '职业'
sheetList = [['小石头1','28','female'],['xing.org1^','29','male']] # 嵌套list，为输入多行数据准备

for item in sheetList:
    sheet.append(item)
# for循环输入多行数据

wb.save(site + '6-4-人员信息表.xlsx') # 存储excel工作簿
print('数据输入完成')


print('开始读取数据')
#读取的代码：
wb = openpyxl.load_workbook(site+ '6-4-人员信息表.xlsx') # 打开一个工作簿(没有找到会报错：FileNotFoundError: [Errno 2] No such file or directory: 'Marvel.xlsx')
print(wb) # <openpyxl.workbook.workbook.Workbook object at 0x10e423278>
sheet = wb['myOneSheet'] # 获取对应名称的工作表（？如果没有这个名字怎么办？不能实现获取一下sheet的title嘛？）
sheetname = wb.sheetnames # 获取sheet的name
print(sheetname)

A1_cell = sheet['A1'] # 获取第一个单元格对象'<Cell 'new title'.A1>'
A1_value = A1_cell.value # 获取第一个单元格的value值（也就是内容）
print(A1_value)